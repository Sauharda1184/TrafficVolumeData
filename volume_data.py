"""
volume_data.py
Generalizable pipeline for any intersection's bin-statistics data.

Usage
-----
1. Edit the CONFIG section below.
2. Run:  python3 volume_data.py

For each approach + movement found in the data the script produces:
  <OUTPUT_DIR>/<APPROACH>_<MOVEMENT>.csv      — hourly pivot table
  <OUTPUT_DIR>/<APPROACH>_<MOVEMENT>.xlsx     — Excel workbook with chart
"""

import csv
import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — edit this section for each new intersection
# ─────────────────────────────────────────────────────────────────────────────

INTERSECTION = "CSAH 61 (Flying Cloud Dr) at College View Dr"

# Map each approach code to the directory that holds its CSV files.
# Add or remove entries for 2- or 4-legged intersections.
APPROACH_DIRS = {
    "NB": "NB_DATA",
    "SB": "SB_DATA",
    # "EB": "EB_DATA",
    # "WB": "WB_DATA",
}

# Where to write output files (created automatically if it doesn't exist).
OUTPUT_DIR = "output"

# Movement types to process.  T=Through  L=Left  R=Right
# Remove any movements you don't need (e.g. remove "R" to skip right turns).
MOVEMENTS = ["T", "L", "R"]

MOVEMENT_LABELS = {
    "T": "Through",
    "L": "Left Turn",
    "R": "Right Turn",
}

APPROACH_LABELS = {
    "NB": "Northbound",
    "SB": "Southbound",
    "EB": "Eastbound",
    "WB": "Westbound",
}

# ─────────────────────────────────────────────────────────────────────────────
# Styling constants (match the existing _GRAPH.png palette)
# ─────────────────────────────────────────────────────────────────────────────

HOURS      = list(range(23))
HOUR_LABELS = [f"{h:02d}:00" for h in HOURS]
INTERVALS   = [f"{h:02d}:{m:02d}" for h in range(23) for m in (0, 15, 30, 45)]

DAY_COLORS = [
    "e6194b", "3cb44b", "4363d8", "f58231", "911eb4",
    "42d4f4", "f032e6", "a3c832", "f4a8c0", "469990",
]


# Columns used when a zone covers multiple movements (combo zones like WBTR1)
COMBO_COLUMNS = {
    "T": "ThroughCount",
    "L": "LeftTurnCount",
    "R": "RightTurnCount",
}

# ─────────────────────────────────────────────────────────────────────────────
# Data processing
# ─────────────────────────────────────────────────────────────────────────────

def scan_all_zones(directory):
    """Return sorted list of every unique ZoneName found across all CSVs in directory."""
    zones = set()
    for fname in sorted(Path(directory).glob("*.csv")):
        with open(fname, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                z = row.get("ZoneName", "").strip()
                if z:
                    zones.add(z)
    return sorted(zones)


def config_to_zone_map(zone_config):
    """
    Convert GUI zone config to the zone_map format used by process_file().

    zone_config: {zone_name: {"T": col_or_None, "L": col_or_None, "R": col_or_None}}
    Returns:     {zone_name: [(movement, column), ...]}   (only non-Skip entries)
    """
    zone_map = {}
    for zone, mv_cols in zone_config.items():
        mappings = [(mv, col) for mv, col in mv_cols.items() if col and col != "Skip"]
        if mappings:
            zone_map[zone] = mappings
    return zone_map


def discover_zones(directory, approach):
    """
    Scan all CSV files in *directory* and return a zone map:
      { zone_name: [(movement, column), ...] }

    Standard single-movement zone  → uses Volume column
      e.g. "NBT1"  → [("T", "Volume")]
           "NBL1"  → [("L", "Volume")]

    Combo multi-movement zone → uses specific count columns
      e.g. "WBTR1" → [("T", "ThroughCount"), ("R", "RightTurnCount")]
           "NBTL2" → [("T", "ThroughCount"), ("L", "LeftTurnCount")]
    """
    zone_pattern = re.compile(
        rf"^{re.escape(approach)}([TLRtlr]+)(\d+)$", re.IGNORECASE
    )
    zone_map = {}

    for fname in sorted(Path(directory).glob("*.csv")):
        with open(fname, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                zone = row.get("ZoneName", "").strip()
                if zone in zone_map:
                    continue
                m = zone_pattern.match(zone)
                if not m:
                    continue
                letters   = m.group(1).upper()
                movements = [l for l in letters if l in MOVEMENTS]
                if not movements:
                    continue
                if len(movements) == 1:
                    # Standard zone — Volume covers the whole movement
                    zone_map[zone] = [(movements[0], "Volume")]
                else:
                    # Combo zone — each movement has its own count column
                    zone_map[zone] = [(mv, COMBO_COLUMNS[mv]) for mv in movements]

    return zone_map


def day_label(filename):
    """
    Convert a filename to a readable day label.
    E.g. 'NB_June_1.csv' -> 'June-1'
         'NB_MAY_4.csv'  -> 'May-4'
    """
    stem  = Path(filename).stem
    parts = stem.split("_")
    return f"{parts[-2].capitalize()}-{parts[-1]}"


def process_file(filepath, zone_map):
    """
    Read one CSV file and return hourly totals per movement.
    zone_map: { zone_name: [(movement, column), ...] }

    Returns: {movement: {hour: total_volume}}
    """
    hourly = defaultdict(lambda: defaultdict(int))

    with open(filepath, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            zone = row.get("ZoneName", "").strip()
            if zone not in zone_map:
                continue
            ts   = row["TimeStamp"].strip()
            hour = int(ts[11:13])
            for movement, column in zone_map[zone]:
                value = int(float(row.get(column, 0) or 0))
                hourly[movement][hour] += value

    return {mv: dict(hours) for mv, hours in hourly.items()}


def load_all_files(directory, zone_map):
    """
    Process every CSV in *directory* and return pivoted data per movement.

    Returns: {movement: {day_label: {hour: volume}}}
    """
    pivot = defaultdict(dict)

    for fname in sorted(Path(directory).glob("*.csv"),
                        key=lambda p: [int(x) if x.isdigit() else x
                                       for x in re.split(r"(\d+)", p.stem)]):
        label     = day_label(fname.name)
        file_data = process_file(fname, zone_map)
        for movement, hourly in file_data.items():
            pivot[movement][label] = hourly

    return dict(pivot)


def process_file_15min(filepath, zone_map):
    """
    Read one CSV file and return raw 15-minute interval totals per movement.
    Returns: {movement: {"HH:MM": total_volume}}
    """
    intervals = defaultdict(lambda: defaultdict(int))

    with open(filepath, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            zone = row.get("ZoneName", "").strip()
            if zone not in zone_map:
                continue
            ts       = row["TimeStamp"].strip()
            interval = ts[11:16]              # "HH:MM"
            for movement, column in zone_map[zone]:
                value = int(float(row.get(column, 0) or 0))
                intervals[movement][interval] += value

    return {mv: dict(ivs) for mv, ivs in intervals.items()}


def load_all_files_15min(directory, zone_map):
    """
    Process every CSV in *directory* and return pivoted 15-min data per movement.
    Returns: {movement: {day_label: {"HH:MM": volume}}}
    """
    pivot = defaultdict(dict)

    for fname in sorted(Path(directory).glob("*.csv"),
                        key=lambda p: [int(x) if x.isdigit() else x
                                       for x in re.split(r"(\d+)", p.stem)]):
        label     = day_label(fname.name)
        file_data = process_file_15min(fname, zone_map)
        for movement, iv_data in file_data.items():
            pivot[movement][label] = iv_data

    return dict(pivot)


# ─────────────────────────────────────────────────────────────────────────────
# CSV output
# ─────────────────────────────────────────────────────────────────────────────

def write_csv(pivot_data, days, col_name, out_path):
    """Write hourly pivot table + TOTAL row to a CSV file."""
    n = len(days)
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Hour"] + days + ["Average"])
        for hour, label in zip(HOURS, HOUR_LABELS):
            vals = [pivot_data.get(d, {}).get(hour, 0) for d in days]
            avg  = round(sum(vals) / n, 1) if n else 0
            writer.writerow([label] + vals + [avg])
        totals = [sum(pivot_data.get(d, {}).get(h, 0) for h in HOURS) for d in days]
        avg_total = round(sum(totals) / n, 1) if n else 0
        writer.writerow(["Total"] + totals + [avg_total])
    print(f"  CSV  → {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(pivot_data, days, title, out_path, pivot_15min=None):
    """Build a styled Excel workbook with a Data sheet and a Chart sheet.
    If pivot_15min is provided, also adds '15-min Data' and '15-min Chart' sheets."""
    wb = Workbook()

    # ── Data sheet ───────────────────────────────────────────────────────────
    ws        = wb.active
    ws.title  = "Data"
    n_days    = len(days)

    avg_col   = n_days + 2          # column index of the Average column
    avg_fill  = PatternFill("solid", fgColor="404040")
    avg_hfont = Font(bold=True, color="FFFFFF", size=10)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=avg_col)
    tc           = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill      = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[1].height = 24

    # Header row
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    def header_cell(row, col, value, fill=None, font=None):
        c           = ws.cell(row=row, column=col, value=value)
        c.font      = font or hdr_font
        c.fill      = fill or hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border    = _thin_border()

    header_cell(2, 1, "Hour")
    for c, day in enumerate(days, start=2):
        header_cell(2, c, day)
    header_cell(2, avg_col, "Average", fill=avg_fill, font=avg_hfont)

    # Data rows
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    avg_dfill = PatternFill("solid", fgColor="E8E8E8")
    avg_dalt  = PatternFill("solid", fgColor="D8D8D8")
    avg_dfont = Font(bold=True, size=10, color="202020")
    for r, (hour, label) in enumerate(zip(HOURS, HOUR_LABELS), start=3):
        hc           = ws.cell(row=r, column=1, value=label)
        hc.font      = Font(bold=True, size=10)
        hc.alignment = Alignment(horizontal="center")
        hc.border    = _thin_border()
        if r % 2 == 0:
            hc.fill = alt_fill
        vals = []
        for c, day in enumerate(days, start=2):
            val            = pivot_data.get(day, {}).get(hour, 0)
            vals.append(val)
            cell           = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _thin_border()
            if r % 2 == 0:
                cell.fill = alt_fill
        avg_val  = round(sum(vals) / n_days, 1) if n_days else 0
        ac        = ws.cell(row=r, column=avg_col, value=avg_val)
        ac.font   = avg_dfont
        ac.fill   = avg_dalt if r % 2 == 0 else avg_dfill
        ac.alignment = Alignment(horizontal="center")
        ac.border = _thin_border()

    # Total row
    total_row  = 3 + len(HOURS)
    tot_fill   = PatternFill("solid", fgColor="1F4E79")
    tot_font   = Font(bold=True, color="FFFFFF", size=10)
    tc2           = ws.cell(row=total_row, column=1, value="Total")
    tc2.font      = tot_font
    tc2.fill      = tot_fill
    tc2.alignment = Alignment(horizontal="center")
    tc2.border    = _thin_border()
    day_totals = []
    for c, day in enumerate(days, start=2):
        day_total      = sum(pivot_data.get(day, {}).get(h, 0) for h in HOURS)
        day_totals.append(day_total)
        cell           = ws.cell(row=total_row, column=c, value=day_total)
        cell.font      = tot_font
        cell.fill      = tot_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()
    avg_total      = round(sum(day_totals) / n_days, 1) if n_days else 0
    atc            = ws.cell(row=total_row, column=avg_col, value=avg_total)
    atc.font       = Font(bold=True, color="FFFFFF", size=10)
    atc.fill       = PatternFill("solid", fgColor="404040")
    atc.alignment  = Alignment(horizontal="center")
    atc.border     = _thin_border()

    # Column widths + freeze
    ws.column_dimensions["A"].width = 9
    for c in range(2, avg_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B3"

    # ── Chart sheet ──────────────────────────────────────────────────────────
    wc       = wb.create_sheet("Chart")
    chart    = LineChart()
    chart.title        = title
    chart.y_axis.title = "Traffic Volume (vehicles)"
    chart.x_axis.title = "Hour of Day"
    chart.width        = 28
    chart.height       = 15
    chart.y_axis.numFmt = "0"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    data_ref = Reference(ws, min_col=2, max_col=avg_col,
                         min_row=2, max_row=2 + len(HOURS))
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(HOURS))
    chart.set_categories(cats)

    for i, series in enumerate(chart.series):
        if i < n_days:
            color = DAY_COLORS[i % len(DAY_COLORS)]
            series.smooth = True
            series.graphicalProperties.line.solidFill        = color
            series.graphicalProperties.line.width            = 18000
            series.marker.symbol                             = "circle"
            series.marker.size                               = 4
            series.marker.graphicalProperties.solidFill      = color
            series.marker.graphicalProperties.line.solidFill = color
        else:
            # Average — thick dark dashed line, no markers
            series.smooth = True
            series.graphicalProperties.line.solidFill  = "404040"
            series.graphicalProperties.line.width      = 28000
            series.graphicalProperties.line.dashStyle  = "dash"
            series.marker.symbol                       = "none"

    legend          = Legend()
    legend.position = "r"
    legend.overlay  = False
    chart.legend    = legend

    wc.add_chart(chart, "B2")

    # ── 15-min sheets (optional) ─────────────────────────────────────────────
    if pivot_15min is not None:
        _build_15min_data_sheet(wb, pivot_15min, days, title, n_days)
        _build_15min_chart_sheet(wb, pivot_15min, days, title, n_days)

    wb.save(out_path)
    print(f"  XLSX → {out_path}")


def _build_15min_data_sheet(wb, pivot_15min, days, title, n_days):
    """Grouped/expandable sheet: hourly summary rows + collapsible 15-min detail rows."""
    ws      = wb.create_sheet("15-min Data")
    avg_col = n_days + 2

    # Summary rows appear ABOVE their detail rows so [+] sits on the summary row
    ws.sheet_properties.outlinePr.summaryBelow  = False
    ws.sheet_properties.outlinePr.summaryRight  = False

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=avg_col)
    tc           = ws.cell(row=1, column=1, value=f"{title} — 15-Minute Intervals")
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill      = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[1].height = 24

    # Instruction row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=avg_col)
    ic           = ws.cell(row=2, column=1,
                           value="Click [+] on the left margin to expand an hour into its four 15-minute intervals")
    ic.font      = Font(size=9, italic=True, color="444444")
    ic.alignment = Alignment(horizontal="center")
    ic.fill      = PatternFill("solid", fgColor="EBF3FB")
    ws.row_dimensions[2].height = 16

    # Header row
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    avg_hfill = PatternFill("solid", fgColor="404040")
    h = ws.cell(row=3, column=1, value="Hour / Interval")
    h.font = hdr_font; h.fill = hdr_fill
    h.alignment = Alignment(horizontal="center"); h.border = _thin_border()
    for c, day in enumerate(days, start=2):
        cell = ws.cell(row=3, column=c, value=day)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
    ahdr = ws.cell(row=3, column=avg_col, value="Average")
    ahdr.font = Font(bold=True, color="FFFFFF", size=10)
    ahdr.fill = avg_hfill
    ahdr.alignment = Alignment(horizontal="center"); ahdr.border = _thin_border()

    sum_fill   = PatternFill("solid", fgColor="D6E4F0")
    sum_font   = Font(bold=True, size=10)
    det_font   = Font(size=9)
    det_fill_a = PatternFill("solid", fgColor="FFFFFF")
    det_fill_b = PatternFill("solid", fgColor="F5F9FD")
    tot_fill   = PatternFill("solid", fgColor="1F4E79")
    tot_font   = Font(bold=True, color="FFFFFF", size=10)
    avg_sfill  = PatternFill("solid", fgColor="E0E0E0")   # summary average cell
    avg_dfill  = PatternFill("solid", fgColor="F0F0F0")   # detail average cell
    avg_sfont  = Font(bold=True, size=10, color="202020")
    avg_dfont  = Font(size=9, color="202020")

    r = 4
    for h_idx, hour in enumerate(HOURS):
        # ── Summary row (hourly total + average) ─────────────────────────────
        sc = ws.cell(row=r, column=1, value=f"{hour:02d}:00")
        sc.font = sum_font; sc.fill = sum_fill
        sc.alignment = Alignment(horizontal="center"); sc.border = _thin_border()
        hour_vals = []
        for c, day in enumerate(days, start=2):
            val = sum(pivot_15min.get(day, {}).get(f"{hour:02d}:{m:02d}", 0)
                      for m in (0, 15, 30, 45))
            hour_vals.append(val)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = sum_font; cell.fill = sum_fill
            cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
        hour_avg = round(sum(hour_vals) / n_days, 1) if n_days else 0
        sac = ws.cell(row=r, column=avg_col, value=hour_avg)
        sac.font = avg_sfont; sac.fill = avg_sfill
        sac.alignment = Alignment(horizontal="center"); sac.border = _thin_border()
        r += 1

        # ── Detail rows (collapsed by default) ───────────────────────────────
        for i, m in enumerate((0, 15, 30, 45)):
            interval = f"{hour:02d}:{m:02d}"
            fill = det_fill_b if i % 2 else det_fill_a
            dc = ws.cell(row=r, column=1, value=interval)
            dc.font = det_font; dc.fill = fill
            dc.alignment = Alignment(horizontal="center", indent=2); dc.border = _thin_border()
            iv_vals = []
            for c, day in enumerate(days, start=2):
                val  = pivot_15min.get(day, {}).get(interval, 0)
                iv_vals.append(val)
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = det_font; cell.fill = fill
                cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
            iv_avg = round(sum(iv_vals) / n_days, 1) if n_days else 0
            dac = ws.cell(row=r, column=avg_col, value=iv_avg)
            dac.font = avg_dfont; dac.fill = avg_dfill
            dac.alignment = Alignment(horizontal="center"); dac.border = _thin_border()
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden        = True
            r += 1

    # Total row
    tc2 = ws.cell(row=r, column=1, value="Total")
    tc2.font = tot_font; tc2.fill = tot_fill
    tc2.alignment = Alignment(horizontal="center"); tc2.border = _thin_border()
    tot_vals = []
    for c, day in enumerate(days, start=2):
        val  = sum(pivot_15min.get(day, {}).get(iv, 0) for iv in INTERVALS)
        tot_vals.append(val)
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = tot_font; cell.fill = tot_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = _thin_border()
    tot_avg = round(sum(tot_vals) / n_days, 1) if n_days else 0
    tac = ws.cell(row=r, column=avg_col, value=tot_avg)
    tac.font = Font(bold=True, color="FFFFFF", size=10)
    tac.fill = PatternFill("solid", fgColor="404040")
    tac.alignment = Alignment(horizontal="center"); tac.border = _thin_border()

    ws.column_dimensions["A"].width = 15
    for c in range(2, avg_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B4"


def _build_15min_chart_sheet(wb, pivot_15min, days, title, n_days):
    """Flat 92-row data table + line chart on a dedicated sheet."""
    wc      = wb.create_sheet("15-min Chart")
    avg_col = n_days + 2

    # ── Data table (chart draws from this) ───────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="2E75B6")
    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    avg_hfill = PatternFill("solid", fgColor="404040")

    h = wc.cell(row=1, column=1, value="Interval")
    h.font = hdr_font; h.fill = hdr_fill
    h.alignment = Alignment(horizontal="center")
    for c, day in enumerate(days, start=2):
        cell = wc.cell(row=1, column=c, value=day)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ahdr = wc.cell(row=1, column=avg_col, value="Average")
    ahdr.font = Font(bold=True, color="FFFFFF", size=9)
    ahdr.fill = avg_hfill
    ahdr.alignment = Alignment(horizontal="center")

    for r, interval in enumerate(INTERVALS, start=2):
        wc.cell(row=r, column=1, value=interval)
        iv_vals = []
        for c, day in enumerate(days, start=2):
            val = pivot_15min.get(day, {}).get(interval, 0)
            iv_vals.append(val)
            wc.cell(row=r, column=c, value=val)
        avg_val = round(sum(iv_vals) / n_days, 1) if n_days else 0
        wc.cell(row=r, column=avg_col, value=avg_val)

    # ── Chart ─────────────────────────────────────────────────────────────────
    chart              = LineChart()
    chart.title        = f"{title} — 15-Minute Intervals"
    chart.y_axis.title = "Traffic Volume (vehicles)"
    chart.x_axis.title = "15-Minute Interval"
    chart.width        = 32
    chart.height       = 16
    chart.y_axis.numFmt = "0"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    data_ref = Reference(wc, min_col=2, max_col=avg_col,
                         min_row=1, max_row=1 + len(INTERVALS))
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(wc, min_col=1, min_row=2, max_row=1 + len(INTERVALS))
    chart.set_categories(cats)

    for i, series in enumerate(chart.series):
        if i < n_days:
            color = DAY_COLORS[i % len(DAY_COLORS)]
            series.smooth          = True
            series.graphicalProperties.line.solidFill        = color
            series.graphicalProperties.line.width            = 15000
            series.marker.symbol                             = "circle"
            series.marker.size                               = 3
            series.marker.graphicalProperties.solidFill      = color
            series.marker.graphicalProperties.line.solidFill = color
        else:
            # Average — thick dark dashed line, no markers
            series.smooth = True
            series.graphicalProperties.line.solidFill  = "404040"
            series.graphicalProperties.line.width      = 25000
            series.graphicalProperties.line.dashStyle  = "dash"
            series.marker.symbol                       = "none"

    legend          = Legend()
    legend.position = "r"
    legend.overlay  = False
    chart.legend    = legend

    wc.add_chart(chart, f"{get_column_letter(avg_col + 2)}1")

    wc.column_dimensions["A"].width = 9
    for c in range(2, avg_col + 1):
        wc.column_dimensions[get_column_letter(c)].width = 10


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for approach, directory in APPROACH_DIRS.items():
        print(f"\n{'─'*50}")
        print(f"Approach: {approach}  ({directory})")

        if not os.path.isdir(directory):
            print(f"  [SKIP] Directory not found: {directory}")
            continue

        zone_map = discover_zones(directory, approach)
        if not zone_map:
            print(f"  [SKIP] No matching zones found in {directory}")
            continue

        for zone, mappings in zone_map.items():
            desc = ", ".join(f"{mv}←{col}" for mv, col in mappings)
            print(f"  {zone:12s}  {desc}")

        all_pivot    = load_all_files(directory, zone_map)
        all_pivot_15 = load_all_files_15min(directory, zone_map)

        for movement, pivot_data in all_pivot.items():
            days  = list(pivot_data.keys())
            label = MOVEMENT_LABELS.get(movement, movement)
            approach_label = APPROACH_LABELS.get(approach, approach)
            title = f"{approach_label} {label}\n{INTERSECTION}"
            col   = f"{approach}{movement}"
            stem  = f"{approach}_{movement}"

            csv_path  = os.path.join(OUTPUT_DIR, f"{stem}.csv")
            xlsx_path = os.path.join(OUTPUT_DIR, f"{stem}.xlsx")

            write_csv(pivot_data, days, col, csv_path)
            build_excel(pivot_data, days, title.replace("\n", " — "), xlsx_path,
                        pivot_15min=all_pivot_15.get(movement))

    print(f"\nDone. All files written to '{OUTPUT_DIR}/'")


if __name__ == "__main__":
    main()
