import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

INTERSECTION = "CSAH 61 (Flying Cloud Dr) at College View Dr"

EXPORTS = [
    {
        "csvs":  ["NB_THROUGH_DATA.csv"],
        "col":   "NBT",
        "title": f"Northbound Through\n{INTERSECTION}",
        "out":   "NB_THROUGH_POINTS.xlsx",
    },
    {
        "csvs":  ["NB_LEFT_DATA.csv"],
        "col":   "NBL",
        "title": f"Northbound Left Turn\n{INTERSECTION}",
        "out":   "NB_LEFT_POINTS.xlsx",
    },
    {
        "csvs":  ["SB_THROUGH_DATA.csv"],
        "col":   "SBT",
        "title": f"Southbound Through\n{INTERSECTION}",
        "out":   "SB_THROUGH_POINTS.xlsx",
    },
    {
        "csvs":  ["SB_LEFT_DATA.csv"],
        "col":   "SBL",
        "title": f"Southbound Left Turn\n{INTERSECTION}",
        "out":   "SB_LEFT_POINTS.xlsx",
    },
]

HOURS = list(range(23))
HOUR_LABELS = [f"{h:02d}:00" for h in HOURS]

# Match the exact colors from the matplotlib graphs
DAY_COLORS = [
    "e6194b", "3cb44b", "4363d8", "f58231", "911eb4",
    "42d4f4", "f032e6", "a3c832", "f4a8c0", "469990",
]


def day_label(filename):
    stem  = Path(filename).stem
    parts = stem.split("_")
    return f"{parts[-2].capitalize()}-{parts[-1]}"


def load_pivoted(csv_files, col):
    day_data  = {}
    day_order = []
    for csv_path in csv_files:
        with open(csv_path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["File"] == "TOTAL":
                    continue
                label = day_label(row["File"])
                hour  = int(row["Hour"])
                vol   = int(row[col])
                if label not in day_data:
                    day_data[label] = {}
                    day_order.append(label)
                day_data[label][hour] = vol
    return day_order, day_data


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def build_xlsx(cfg):
    days, data = load_pivoted(cfg["csvs"], cfg["col"])

    wb = Workbook()

    # ── Data sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 1)
    title_cell            = ws.cell(row=1, column=1, value=cfg["title"].replace("\n", " — "))
    title_cell.font       = Font(bold=True, size=13, color="FFFFFF")
    title_cell.alignment  = Alignment(horizontal="center", vertical="center")
    title_cell.fill       = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[1].height = 24

    # Header row
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    h_cell            = ws.cell(row=2, column=1, value="Hour")
    h_cell.font       = header_font
    h_cell.fill       = header_fill
    h_cell.alignment  = Alignment(horizontal="center")
    h_cell.border     = thin_border()

    for c, day in enumerate(days, start=2):
        cell           = ws.cell(row=2, column=c, value=day)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border()

    # Data rows — use "HH:00" labels in column A to match graph x-axis
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for r, (hour, label) in enumerate(zip(HOURS, HOUR_LABELS), start=3):
        hour_cell           = ws.cell(row=r, column=1, value=label)
        hour_cell.font      = Font(bold=True, size=10)
        hour_cell.alignment = Alignment(horizontal="center")
        hour_cell.border    = thin_border()
        if r % 2 == 0:
            hour_cell.fill = alt_fill

        for c, day in enumerate(days, start=2):
            val            = data[day].get(hour, 0)
            cell           = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border()
            if r % 2 == 0:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 9
    for c in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B3"

    # ── Chart sheet ─────────────────────────────────────────────────────────
    wc = wb.create_sheet("Chart")

    chart              = LineChart()
    chart.title        = cfg["title"].replace("\n", "  |  ")
    chart.y_axis.title = "Traffic Volume (vehicles)"
    chart.x_axis.title = "Hour of Day"
    chart.width        = 28       # wider so legend on right doesn't overlap plot
    chart.height       = 15
    chart.y_axis.numFmt        = "0"
    chart.y_axis.crossAx       = 100
    chart.x_axis.crossAx       = 200
    chart.x_axis.tickLblSkip   = 1
    chart.x_axis.tickMarkSkip  = 1

    # Data + categories
    data_ref = Reference(ws, min_col=2, max_col=len(days) + 1,
                         min_row=2, max_row=2 + len(HOURS))
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(HOURS))
    chart.set_categories(cats)

    # Style each series to match DAY_COLORS + add circle markers
    for i, series in enumerate(chart.series):
        color = DAY_COLORS[i % len(DAY_COLORS)]
        series.smooth = True
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width     = 18000   # 1.4 pt
        series.marker.symbol                      = "circle"
        series.marker.size                        = 4
        series.marker.graphicalProperties.solidFill             = color
        series.marker.graphicalProperties.line.solidFill        = color

    # Legend — pinned to the right, overlay off so it never covers the plot
    legend          = Legend()
    legend.position = "r"
    legend.overlay  = False
    chart.legend    = legend

    wc.add_chart(chart, "B2")

    wb.save(cfg["out"])
    print(f"Saved {cfg['out']}")


if __name__ == "__main__":
    for cfg in EXPORTS:
        build_xlsx(cfg)
